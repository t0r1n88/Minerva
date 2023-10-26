"""
Скрипт для генерации некоторых отчетов полезных в работе над билетом в будущее
"""
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import time
from tkinter import messagebox
import re
import warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
warnings.simplefilter(action='ignore', category=DeprecationWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
pd.options.mode.chained_assignment = None


def main_create_report(path_file_data:str,path_end_folder):
    """
    Основной скрипт по созданию отчетов
    :param path_file_data: путь к файлу с данными
    :param path_end_folder: папка куда будут сохранены данные
    :return:
    """

    t = time.localtime()
    current_time = time.strftime('%H_%M_%S', t) # получаем время запуска
    df = pd.read_excel(path_file_data,dtype=str,skiprows=3)
    # Заполняем пустые ячейки
    df['названия'] = df['названия'].fillna('Пусто')
    df['даты прохождения'] = df['даты прохождения'].fillna('Пусто')
    df['темы диагностик'] = df['темы диагностик'].fillna('Пусто')
    df['даты пройденных диагностик'] = df['даты пройденных диагностик'].fillna('Пусто')
    begin = pd.to_datetime('2023-09-01')
    end = pd.to_datetime('2023-12-01')

    wb_more_probs = generate_enroll_more_probs(df.copy(),begin,end)
    wb_diag = generate_report_diag(df.copy(),begin,end)

    wb_more_probs.save(f'{path_end_folder}/Записавшиеся на 2 и более профпроб {current_time}.xlsx')
    wb_diag.save(f'{path_end_folder}/Свод по диагностике {current_time}.xlsx')





def generate_enroll_more_probs(df:pd.DataFrame,begin_event,end_event):
    """
    Функция для генерации списка школьников записавшихся на 2 и более профпроб в 2023 году
    :param df: датафрейм
    :param begin_event: дата начала
    :param end_event: дата конца
    :return:
    """

    wb = openpyxl.Workbook()  # файл для записавшихся на 2 и более профпробы
    wb.create_sheet('Записавшиеся на 2 и более проб', index=0) # создаем лист



    # разворачиваем датафрейм чтобы названия и даты профпроб развернулись в новые строки
    flat_rows = []  # список для хранения строк
    for index, row in df.iterrows():
        name_probs = row['названия'].split(',')  # названия проб
        date_probs = row['даты прохождения'].split(',')  # даты прохождения
        for idx, name in enumerate(name_probs):
            new_row = row.copy()  # копируем строку
            new_row['названия'] = name
            try:
                new_row['даты прохождения'] = date_probs[idx]
                flat_rows.append(new_row)
            except IndexError:
                new_row['даты прохождения'] = date_probs[idx - 1]
                flat_rows.append(new_row)
    # Создание нового датафрейма на основе новых строк
    flat_df = pd.DataFrame(flat_rows)
    # Приводим к типу дата
    flat_df['даты прохождения'] = flat_df['даты прохождения'].astype(str)
    flat_df['даты прохождения'] = flat_df['даты прохождения'].apply(lambda x: x.strip())
    flat_df = flat_df[flat_df['даты прохождения'] != 'Пусто'] # удаляем тех кто не проходил
    flat_df['даты прохождения'] = pd.to_datetime(flat_df['даты прохождения'],
                                                 errors='ignore')  # приводим к типу дата

    flat_df = flat_df[(flat_df['даты прохождения'] >= begin_event) & (
                flat_df['даты прохождения'] <= end_event)]  # отбираем только те профпробы которые были этой осенью

    dupl_df = flat_df[flat_df.duplicated(subset=['ФИО', 'Дата рождения'], keep=False)]  # отбираем дубликаты
    lst_out = ['Муниципалитет', 'ФИО', 'Дата рождения', 'образовательная организация', 'класс (без буквы)',
               'буква класса',
               'названия', 'даты прохождения', 'ответственный педагог-навигатор']

    dupl_df = dupl_df[lst_out]

    dupl_df = dupl_df.sort_values(by=['ФИО', 'даты прохождения'], ascending=True)
    dupl_df['даты прохождения'] = dupl_df['даты прохождения'].dt.strftime('%d.%m.%Y')



    for row in dataframe_to_rows(dupl_df, index=False, header=True):
        wb['Записавшиеся на 2 и более проб'].append(row)
    for column in wb['Записавшиеся на 2 и более проб'].columns:
        max_length = 0
        column_name = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        wb['Записавшиеся на 2 и более проб'].column_dimensions[column_name].width = adjusted_width
    return wb


def generate_report_diag(diag_df:pd.DataFrame,begin_event,end_event)->openpyxl.Workbook:
    """
    Функция для генерации количества школьников прошедщих хотя бы одну диагностику по муниципалитетам и школам
    :param diag_df: датафрейм
    :param begin_event:дата начала
    :param begin_event:дата конца
    :return: файл openpyxl
    """
    # разворачиваем датафрейм чтобы названия и даты диагностик записанные через запятую создали новые строки
    diag_rows = []  # список для хранения строк
    for index, row in diag_df.iterrows():
        name_diag = row['темы диагностик'].split(',')  # названия проб
        date_diag = row['даты пройденных диагностик'].split(',')  # даты прохождения
        for idx, name in enumerate(name_diag):
            new_row = row.copy()  # копируем строку
            new_row['темы диагностик'] = name
            try:
                new_row['даты пройденных диагностик'] = date_diag[idx]
                diag_rows.append(new_row)
            except IndexError:
                new_row['даты пройденных диагностик'] = date_diag[-1]
                diag_rows.append(new_row)
    # Создание нового датафрейма на основе новых строк
    diag_df = pd.DataFrame(diag_rows)
    diag_df['даты пройденных диагностик'] = diag_df['даты пройденных диагностик'].astype(str)
    diag_df['даты пройденных диагностик'] = diag_df['даты пройденных диагностик'].apply(lambda x: x.strip())
    diag_df = diag_df[diag_df['даты пройденных диагностик'] != 'Пусто']  # удаляем тех кто не проходил
    diag_df['даты пройденных диагностик'] = pd.to_datetime(diag_df['даты пройденных диагностик'],
                                                           errors='ignore')  # приводим к типу дата

    diag_df = diag_df[(diag_df['даты пройденных диагностик'] >= begin_event) & (diag_df[
                                                                                    'даты пройденных диагностик'] <= end_event)]  # отбираем только те профпробы которые были этой осенью


    drop_dupl_diag_df = diag_df.drop_duplicates(subset=['ФИО', 'Дата рождения', 'образовательная организация']) # убираем лишние строки
    # делаем сводную таблицу
    svod_diag_df = pd.pivot_table(drop_dupl_diag_df,
                                  index=['Муниципалитет', 'образовательная организация'],
                                  values='ФИО',
                                  aggfunc='count',
                                  margins=True)

    svod_diag_df = svod_diag_df.reset_index()  # вытаскиваеем индекс
    svod_diag_df.rename(columns={'ФИО': 'Количество прошедших хотя бы 1 диагностику'}, inplace=True)
    svod_diag_df.sort_values(by=['Муниципалитет', 'Количество прошедших хотя бы 1 диагностику'],
                             ascending=False, inplace=True)
    lst_sheet = svod_diag_df['Муниципалитет'].unique()  # список значений

    wb_diag = openpyxl.Workbook()  # создаем файл
    wb_diag.create_sheet(title='Общий свод', index=0)
    for row in dataframe_to_rows(svod_diag_df, index=False, header=True):
        wb_diag['Общий свод'].append(row)

    # Устанавливаем автоширину для каждой колонки
    for column in wb_diag['Общий свод'].columns:
        max_length = 0
        column_name = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        wb_diag['Общий свод'].column_dimensions[column_name].width = adjusted_width

    used_name_sheet = set()  # множество для хранения значений которые уже были использованы
    name_column = 'Муниципалитет'
    for idx, value in enumerate(lst_sheet, 1):
        temp_df = svod_diag_df[svod_diag_df[name_column] == value]  # отфильтровываем по значению
        short_value = value[:20]  # получаем обрезанное значение
        short_value = re.sub(r'[\[\]\'+()<> :"?*|\\/]', '_', short_value)

        if short_value in used_name_sheet:
            short_value = f'{short_value}_{idx}'  # добавляем окончание
        wb_diag.create_sheet(short_value, index=idx)  # создаем лист
        used_name_sheet.add(short_value)
        # создаем строку Сумма
        temp_df.loc['Сумма'] = temp_df['Количество прошедших хотя бы 1 диагностику'].sum()
        temp_df.loc[temp_df.index[-1], 'Муниципалитет'] = None
        temp_df.loc[temp_df.index[-1], 'образовательная организация'] = None
        for row in dataframe_to_rows(temp_df, index=False, header=True):
            wb_diag[short_value].append(row)

        # Устанавливаем автоширину для каждой колонки
        for column in wb_diag[short_value].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            wb_diag[short_value].column_dimensions[column_name].width = adjusted_width

    wb_diag.create_sheet()

    if 'All' and 'Sheet' in wb_diag.sheetnames:  # удаляем лишние листы
        del wb_diag['All']
        del wb_diag['Sheet']

    return wb_diag


if __name__=='__main__':
    path_main = 'data/Генерация отчетов/students_all.xlsx'
    path_end_main = 'data/Генерация отчетов/result'
    main_create_report(path_main,path_end_main)

    print('Lindy Booth!')




